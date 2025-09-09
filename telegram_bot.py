import os
import re
import logging
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import datetime
import pytz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# --- Basic Logging Setup ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Configuration ---
# It is recommended to use environment variables for sensitive data
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "YOUR_TELEGRAM_BOT_TOKEN")
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "excelmerge")
CAMBODIA_TZ = pytz.timezone('Asia/Phnom_Penh')

# --- Data Storage (in-memory, resets on restart) ---
user_data = {}
user_breaks = {}

# --- Helper Functions ---
def get_now():
    """Gets the current time in Cambodia timezone."""
    return datetime.datetime.now(CAMBODIA_TZ)

# --- Error Handler ---
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Logs the error and sends a telegram message to notify the developer."""
    logger.error("Exception while handling an update:", exc_info=context.error)


# --- Command Handlers ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Sends a welcome message when the /start command is issued."""
    await update.message.reply_text(
        "Welcome to the Time Tracker Bot!\n\n"
        "You can use the following commands:\n"
        "- `check in`\n"
        "- `check out`\n"
        "- `wc`\n"
        "- `smoke`\n"
        "- `eat`\n\n"
        "When you are back from a break, please reply with `1`, `+1`, `back`, `finish`, or `done`."
    )

async def check_in(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'check in' message."""
    user = update.message.from_user
    user_id = user.id
    now = get_now()
    
    if user_id not in user_data:
        user_data[user_id] = {
            "name": user.full_name,
            "check_in": None,
            "check_out": None,
            "wc_count": 0, "wc_late": 0,
            "smoke_count": 0, "smoke_late": 0,
            "eat_count": 0, "eat_late": 0
        }
    
    user_data[user_id]["check_in"] = now

    # Check-in time logic
    if now.time() >= datetime.time(16, 0) and now.time() < datetime.time(17, 0):
        await update.message.reply_text("Well done!")
    elif now.time() > datetime.time(17, 9) and now.time() < datetime.time(20, 0):
        late_minutes = int((now - now.replace(hour=17, minute=0, second=0, microsecond=0)).total_seconds() / 60)
        await update.message.reply_text(f"You are late by {late_minutes} minutes.")
    elif now.time() > datetime.time(23, 1) and now.time() < datetime.time(23, 59):
        late_minutes = int((now - now.replace(hour=23, minute=0, second=0, microsecond=0)).total_seconds() / 60)
        await update.message.reply_text(f"You are late by {late_minutes} minutes (from 11 PM).")


async def check_out(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'check out' message."""
    user_id = update.message.from_user.id
    now = get_now()

    # Define valid check-out times
    checkout_break_start = datetime.time(23, 0)
    checkout_break_end = datetime.time(23, 59, 59)
    checkout_final_start = datetime.time(5, 0)
    checkout_final_end = datetime.time(6, 0)

    is_valid_time = (now.time() >= checkout_break_start and now.time() <= checkout_break_end) or \
                    (now.time() >= checkout_final_start and now.time() <= checkout_final_end)

    if is_valid_time:
        if user_id in user_data:
            user_data[user_id]["check_out"] = now
        # No message is sent for a valid checkout, as per user preference.
    else:
        await update.message.reply_text("You are not allowed to check out at this time.")


async def wc(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'wc' message."""
    user = update.message.from_user
    user_id = user.id
    
    # If user has no record, create one
    if user_id not in user_data:
        user_data[user_id] = {
            "name": user.full_name,
            "check_in": None,
            "check_out": None,
            "wc_count": 0, "wc_late": 0,
            "smoke_count": 0, "smoke_late": 0,
            "eat_count": 0, "eat_late": 0
        }

    if user_id not in user_breaks:
        user_breaks[user_id] = {"type": "wc", "start_time": get_now()}


async def smoke(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'smoke' message."""
    user = update.message.from_user
    user_id = user.id

    # If user has no record, create one
    if user_id not in user_data:
        user_data[user_id] = {
            "name": user.full_name,
            "check_in": None,
            "check_out": None,
            "wc_count": 0, "wc_late": 0,
            "smoke_count": 0, "smoke_late": 0,
            "eat_count": 0, "eat_late": 0
        }

    if user_id not in user_breaks:
        user_breaks[user_id] = {"type": "smoke", "start_time": get_now()}


async def eat(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'eat' message."""
    user = update.message.from_user
    user_id = user.id
    now = get_now()

    eat_time1_start = now.replace(hour=21, minute=0, second=0, microsecond=0)
    eat_time1_end = now.replace(hour=21, minute=30, second=0, microsecond=0)
    eat_time2_start = now.replace(hour=1, minute=0, second=0, microsecond=0)
    eat_time2_end = now.replace(hour=1, minute=30, second=0, microsecond=0)

    if (now >= eat_time1_start and now <= eat_time1_end) or \
       (now >= eat_time2_start and now <= eat_time2_end):
        
        # If user has no record, create one
        if user_id not in user_data:
            user_data[user_id] = {
                "name": user.full_name,
                "check_in": None,
                "check_out": None,
                "wc_count": 0, "wc_late": 0,
                "smoke_count": 0, "smoke_late": 0,
                "eat_count": 0, "eat_late": 0
            }

        if user_id not in user_breaks:
            user_breaks[user_id] = {"type": "eat", "start_time": get_now()}
    else:
        await update.message.reply_text("It's not time to eat yet.")

async def back_from_break(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles replies for returning from breaks."""
    user_id = update.message.from_user.id
    if user_id in user_breaks:
        break_info = user_breaks.pop(user_id)
        break_type = break_info["type"]
        start_time = break_info["start_time"]
        end_time = get_now()
        duration = int((end_time - start_time).total_seconds() / 60)

        if user_id in user_data:
            if break_type == 'wc':
                user_data[user_id]["wc_count"] += 1
                if duration > 10:
                    late_minutes = duration - 10
                    user_data[user_id]["wc_late"] += late_minutes
                    await update.message.reply_text(f"You are late {late_minutes} minutes.")
            elif break_type == 'smoke':
                user_data[user_id]["smoke_count"] += 1
                if duration > 5:
                    late_minutes = duration - 5
                    user_data[user_id]["smoke_late"] += late_minutes
                    await update.message.reply_text(f"You are late {late_minutes} minutes.")
            elif break_type == 'eat':
                user_data[user_id]["eat_count"] += 1
                late_minutes = 0
                eat_time1_end = end_time.replace(hour=21, minute=30, second=0, microsecond=0)
                eat_time2_end = end_time.replace(hour=1, minute=30, second=0, microsecond=0)
                if start_time.hour >= 21 and end_time > eat_time1_end:
                    late_minutes = int((end_time - eat_time1_end).total_seconds() / 60)
                elif start_time.hour >= 1 and end_time > eat_time2_end:
                    late_minutes = int((end_time - eat_time2_end).total_seconds() / 60)
                
                if late_minutes > 0:
                    user_data[user_id]["eat_late"] += late_minutes
                    await update.message.reply_text(f"You are late {late_minutes} minutes.")


async def get_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generates and sends the daily report if requested by an admin."""
    user = update.message.from_user
    if user.username != ADMIN_USERNAME:
        await update.message.reply_text("You are not authorized to perform this action.")
        return

    if not user_data:
        await update.message.reply_text("No activity to report for today.")
        return

    # Sort users by check-in time, handling cases where check-in is None
    sorted_users = sorted(user_data.items(), key=lambda item: item[1].get('check_in') or datetime.datetime.max.replace(tzinfo=CAMBODIA_TZ))

    report_data = []
    for user_id, data in sorted_users:
        check_in_time = data["check_in"].strftime("%H:%M") if data.get("check_in") else ""
        check_out_time = data["check_out"].strftime("%H:%M") if data.get("check_out") else ""
        
        report_data.append({
            "User": data["name"],
            "Check-in": check_in_time,
            "Check-out": check_out_time,
            "WC": data["wc_count"],
            "WC late (m)": data["wc_late"],
            "Smoke": data["smoke_count"],
            "Smoke late (m)": data["smoke_late"],
            "Eat": data["eat_count"],
            "Eat late (m)": data["eat_late"],
        })
    
    df = pd.DataFrame(report_data)
    file_path = f"daily_report_{get_now().strftime('%Y-%m-%d')}.xlsx"
    df.to_excel(file_path, index=False)
    
    # --- Apply styling to the Excel file ---
    wb = load_workbook(file_path)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="4AACC5", end_color="4AACC5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    late_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    # Find the columns for conditional formatting by header name
    late_columns_indices = []
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value in ["WC late (m)", "Smoke late (m)", "Eat late (m)"]:
            late_columns_indices.append(col_idx)

    # Style the data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
            # Apply conditional formatting for 'late' columns
            if cell.column in late_columns_indices and cell.value is not None:
                try:
                    if int(cell.value) > 0:
                        cell.fill = late_fill
                        cell.value = f"{cell.value} minute"
                except (ValueError, TypeError):
                    pass

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the styled workbook
    wb.save(file_path)

    await context.bot.send_document(chat_id=update.message.chat_id, document=open(file_path, 'rb'))
    os.remove(file_path)

    await update.message.reply_text("Report sent.")

async def clear_data_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Clears all user data for the new day."""
    user_data.clear()
    user_breaks.clear()
    logger.info("Daily user data has been cleared automatically.")

def main() -> None:
    """Start the bot."""
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    # Add the error handler
    application.add_error_handler(error_handler)
    
    # Schedule the daily data clear job
    job_queue = application.job_queue
    clear_time = datetime.time(hour=5, minute=30, tzinfo=CAMBODIA_TZ)
    job_queue.run_daily(clear_data_job, clear_time)


    # on different commands - answer in Telegram
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("getreport", get_report))

    # on non command i.e message - echo the message on Telegram
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^check in$', re.IGNORECASE)), check_in))
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^check out$', re.IGNORECASE)), check_out))
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^wc$', re.IGNORECASE)), wc))
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^smoke$', re.IGNORECASE)), smoke))
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^eat$', re.IGNORECASE)), eat))
    application.add_handler(MessageHandler(filters.Regex(re.compile(r'^(1|\+1|back|finish|done|back to seat)$', re.IGNORECASE)), back_from_break))
    
    # Run the bot until the user presses Ctrl-C
    application.run_polling()

if __name__ == '__main__':
    main()

